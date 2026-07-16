#!/usr/bin/env python3
"""Extract gold-standard benchmark cases from paper supplementary text."""

from __future__ import annotations

import json
import re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SUPP_PATH = PROJECT_ROOT / "paper" / "hema_guide_supplementary.txt"
ENTITY_SLUGS_PATH = PROJECT_ROOT / "data" / "entity_slugs.json"
OUT_DIR = PROJECT_ROOT / "data" / "benchmark"
JSONL_PATH = OUT_DIR / "cases.jsonl"
JSON_PATH = OUT_DIR / "cases.json"
XLSX_PATH = OUT_DIR / "cases.xlsx"
CASES_DIR = OUT_DIR / "cases"

EXCEL_COLUMNS = [
    "case_id",
    "cohort",
    "age",
    "ecog",
    "entity_slug",
    "routing_hint",
    "primary_diagnosis",
    "question",
    "tumor_board_decision",
    "raw_vignette",
]

CASE_START = re.compile(
    r"^(?P<id>(?:l_case_\d+|l_case_sim_\d+|ly_case_\d+|ly_case_sim_\d+|"
    r"mm_case_\d+|mm_case_sim_\d+))$"
)
CASE_END = re.compile(
    r"^(?:"
    r"[a-z_]+__ALL|"
    r"#{3,}|"
    r"={10,}\s*(?:gpt-|qwen|claude|gemini|deepseek|llama|mistral|o[0-9]-)"
    r")",
    re.IGNORECASE,
)
BENCHMARK_IDS = (
    {f"l_case_{i}" for i in range(1, 16)}
    | {f"ly_case_{i:02d}" for i in range(1, 16)}
    | {f"mm_case_{i}" for i in range(1, 16)}
)


def parse_cases(text: str) -> dict[str, str]:
    cases: dict[str, str] = {}
    current_id: str | None = None
    current_lines: list[str] = []

    for line in text.splitlines():
        stripped = line.strip()
        start = CASE_START.match(stripped)
        if start:
            if current_id:
                cases[current_id] = "\n".join(current_lines).strip()
            current_id = start.group("id")
            current_lines = []
            continue

        if current_id is None:
            continue

        if CASE_END.match(stripped):
            cases[current_id] = "\n".join(current_lines).strip()
            current_id = None
            current_lines = []
            continue

        current_lines.append(line)

    if current_id:
        cases[current_id] = "\n".join(current_lines).strip()

    return cases

FIELD_RE = {
    "age": re.compile(r"^Age:\s*(.+)$", re.M | re.I),
    "ecog": re.compile(r"^ECOG:\s*(.+)$", re.M | re.I),
}

PRIMARY_RE = re.compile(
    r"^(?:Primary diagnos(?:is|es)|Hematologic primary diagnosis):\s*\n?"
    r"(.*?)(?=\n(?:Secondary diagnoses|Comorbidities|Treatment and course|"
    r"Prognostic factors|Molecular genetic findings|Question for the tumor board:))",
    re.M | re.S | re.I,
)
QUESTION_RE = re.compile(
    r"Question for the tumor board:\s*\n(.+?)(?=\nTumor board decision:)",
    re.S | re.I,
)
DECISION_RE = re.compile(
    r"Tumor board decision:\s*\n(.+?)(?=\n={10,}|\nTREATMENT TIMELINE|\nGene Transcript|\Z)",
    re.S | re.I,
)

MOLECULAR_Q = re.compile(r"pathogenic\s+variant|targeted\s+therapy", re.I)
ADVANCED_Q = re.compile(
    r"relapse|refractor|further\s+therap|further\s+treatment|"
    r"car-?t|salvage|pd\b|progression|allo-?sct\s+is\s+declined",
    re.I,
)

# Keyword → slug fallback when entity_slugs fuzzy match fails
# Order matters: more specific / primary aggressive entities first.
DIAGNOSIS_HINTS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"promyelocytic|apl\b|aml\s*m3", re.I), "aml"),
    (re.compile(r"acute\s+myeloid|\baml\b", re.I), "aml"),
    (re.compile(r"lymphoblastic|b-all|c-all|tlbl|t-all|\ball\b", re.I), "all"),
    (re.compile(r"\bmds\b|myelodysplastic", re.I), "mds"),
    (re.compile(r"\bcml\b|chronic\s+myeloid", re.I), "cml"),
    (re.compile(r"hairy\s+cell|\bhcl\b", re.I), "hcl"),
    (re.compile(r"dlbcl|diffuse\s+large|pmbcl|mediastinal\s+b-cell", re.I), "dlbcl"),
    (re.compile(r"(?<!non-)(?<!non\s)hodgkin", re.I), "chl"),
    (re.compile(r"malt|marginal\s+zone", re.I), "malt"),
    (re.compile(r"follicular\s+lymphoma", re.I), "fl"),
    (re.compile(r"waldenstr", re.I), "waldenstrom"),
    (re.compile(r"\bmgus\b|monoclonal\s+gammopathy\s+of\s+undetermined", re.I), "myeloma"),
    (re.compile(r"poems", re.I), "myeloma"),
    (re.compile(r"plasma\s+cell\s+leukemia|multiple\s+myeloma|\bmm\b|plasmacytoma", re.I), "myeloma"),
]

# Cases whose gold answer clearly exceeds pure first-line flowchart coverage
ROUTING_OVERRIDES: dict[str, str] = {
    "l_case_12": "advanced",
    "l_case_15": "advanced",
    "ly_case_06": "advanced",
    "ly_case_07": "advanced",
    "mm_case_2": "advanced",
    "mm_case_3": "advanced",
    "mm_case_5": "advanced",
    "mm_case_6": "advanced",
}

# Force entity when vignette primary diagnosis is blank / ambiguous
ENTITY_OVERRIDES: dict[str, str] = {
    "ly_case_09": "dlbcl",
}


def _first_group(pattern: re.Pattern[str], text: str) -> str | None:
    m = pattern.search(text)
    if not m:
        return None
    return m.group(1).strip()


def infer_routing_hint(question: str, decision: str, body: str) -> str:
    blob = f"{question}\n{decision}"
    if MOLECULAR_Q.search(question) or "Molecular genetic findings" in body:
        if MOLECULAR_Q.search(question):
            return "molecular"
    if ADVANCED_Q.search(blob):
        return "advanced"
    return "guideline"


def resolve_entity_slug(primary: str, entity_slugs: dict[str, str]) -> str:
    if not primary:
        return "unknown"
    candidates = [primary.strip(), primary.strip().split("\n")[0].strip()]
    for cand in candidates:
        if cand in entity_slugs:
            return entity_slugs[cand]
    for pattern, slug in DIAGNOSIS_HINTS:
        if pattern.search(primary):
            return slug
    return "unknown"


def parse_case_record(case_id: str, body: str, entity_slugs: dict[str, str]) -> dict:
    age_raw = _first_group(FIELD_RE["age"], body)
    ecog_raw = _first_group(FIELD_RE["ecog"], body)
    primary = _first_group(PRIMARY_RE, body) or ""
    # ly_case_09 has empty primary; fall back to scanning diagnosis-like lines
    if not primary.strip():
        for line in body.splitlines():
            if re.search(r"lymphoma|leukemia|myeloma|mds|cml", line, re.I):
                primary = line.strip()
                break

    question = (_first_group(QUESTION_RE, body) or "").strip()
    decision = (_first_group(DECISION_RE, body) or "").strip()
    # Drop trailing variant tables / timelines still attached
    decision = re.split(r"\n={10,}", decision)[0].strip()

    entity_slug = resolve_entity_slug(primary, entity_slugs)
    if case_id in ENTITY_OVERRIDES:
        entity_slug = ENTITY_OVERRIDES[case_id]
    routing_hint = infer_routing_hint(question, decision, body)
    if case_id in ROUTING_OVERRIDES:
        routing_hint = ROUTING_OVERRIDES[case_id]

    age: int | str | None = None
    if age_raw:
        m = re.search(r"\d+", age_raw)
        age = int(m.group()) if m else age_raw

    ecog: int | str | None = None
    if ecog_raw:
        m = re.search(r"\d+", ecog_raw)
        ecog = int(m.group()) if m else ecog_raw

    return {
        "case_id": case_id,
        "age": age,
        "ecog": ecog,
        "primary_diagnosis": primary.strip(),
        "question": question,
        "tumor_board_decision": decision,
        "entity_slug": entity_slug,
        "routing_hint": routing_hint,
        "raw_vignette": body.strip(),
    }


def cohort_of(case_id: str) -> str:
    if case_id.startswith("l_"):
        return "leukemia"
    if case_id.startswith("ly_"):
        return "lymphoma"
    if case_id.startswith("mm_"):
        return "myeloma_pcd"
    return "other"


def write_excel(records: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "benchmark_cases"

    ws.append(EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rec in records:
        row = []
        for col in EXCEL_COLUMNS:
            if col == "cohort":
                row.append(cohort_of(rec["case_id"]))
            else:
                val = rec.get(col)
                row.append("" if val is None else val)
        ws.append(row)

    # Wrap long text columns
    wrap_cols = {"primary_diagnosis", "question", "tumor_board_decision", "raw_vignette"}
    for col_idx, name in enumerate(EXCEL_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        width = {
            "case_id": 14,
            "cohort": 14,
            "age": 8,
            "ecog": 8,
            "entity_slug": 14,
            "routing_hint": 14,
            "primary_diagnosis": 48,
            "question": 36,
            "tumor_board_decision": 48,
            "raw_vignette": 60,
        }.get(name, 20)
        ws.column_dimensions[letter].width = width
        if name in wrap_cols:
            for cell in ws[letter]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    text = SUPP_PATH.read_text(encoding="utf-8")
    cases = parse_cases(text)
    entity_slugs = json.loads(ENTITY_SLUGS_PATH.read_text(encoding="utf-8"))

    missing = sorted(BENCHMARK_IDS - cases.keys())
    if missing:
        raise SystemExit(f"Missing benchmark cases: {missing}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CASES_DIR.mkdir(parents=True, exist_ok=True)

    records: list[dict] = []
    for case_id in sorted(BENCHMARK_IDS, key=lambda x: (x.split("_")[0], x)):
        body = cases.get(case_id, "")
        if not body:
            continue
        rec = parse_case_record(case_id, body, entity_slugs)
        records.append(rec)
        (CASES_DIR / f"{case_id}.json").write_text(
            json.dumps(rec, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    lean_records = [{k: v for k, v in rec.items() if k != "raw_vignette"} for rec in records]

    with JSONL_PATH.open("w", encoding="utf-8") as f:
        for lean in lean_records:
            f.write(json.dumps(lean, ensure_ascii=False) + "\n")

    JSON_PATH.write_text(
        json.dumps(lean_records, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    write_excel(records, XLSX_PATH)

    by_route: dict[str, int] = {}
    by_slug: dict[str, int] = {}
    for rec in records:
        by_route[rec["routing_hint"]] = by_route.get(rec["routing_hint"], 0) + 1
        by_slug[rec["entity_slug"]] = by_slug.get(rec["entity_slug"], 0) + 1

    print(f"Wrote {len(records)} cases to {JSONL_PATH}")
    print(f"Wrote {JSON_PATH}")
    print(f"Wrote {XLSX_PATH}")
    print(f"Per-case JSON under {CASES_DIR}")
    print(f"routing_hint: {dict(sorted(by_route.items()))}")
    print(f"entity_slug: {dict(sorted(by_slug.items()))}")


if __name__ == "__main__":
    main()
